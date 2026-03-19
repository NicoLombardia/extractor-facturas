"""
Extractor de Facturas PDF → Excel
App web con Streamlit — subible a Streamlit Cloud gratis.
"""

import io
import re
import tempfile
import zipfile
from pathlib import Path
from datetime import datetime

import pdfplumber
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
#  LÓGICA DE EXTRACCIÓN
# ══════════════════════════════════════════════════════════════════════

PATRONES = {
    "numero_factura": [
        r"(?:factura|invoice|comprobante)\s*[nN°#:\.]*\s*([A-Z0-9\-\/]+)",
        r"(?:N°|Nro\.?|No\.?)\s*(?:de\s+)?(?:factura)?\s*[:\-]?\s*([A-Z0-9\-\/]+)",
        r"FACTURA\s+([A-Z0-9\-\/]+)",
    ],
    "fecha_emision": [
        r"(?:fecha|date|emisi[oó]n)[:\s]*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})",
        r"(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{4})",
    ],
    "cuit_emisor": [
        r"(?:CUIT|RUT|RFC|NIF)[:\s]*(\d{2}[-\s]?\d{8}[-\s]?\d{1}|\d{11,13})",
    ],
    "emisor": [
        r"(?:raz[oó]n social|empresa|proveedor|emisor)[:\s]*([^\n\r]{3,60})",
        r"^([A-Z][A-Z\s&\.]{5,50}(?:S\.A\.|S\.R\.L\.|SAS|LTDA|INC\.?))",
    ],
    "cliente": [
        r"(?:cliente|receptor|comprador|facturado a)[:\s]*([^\n\r]{3,60})",
        r"(?:señor(?:es)?|sr\.?)[:\s]*([^\n\r]{3,60})",
    ],
    "subtotal": [r"(?:subtotal|base imponible|neto)[:\s$]*([0-9]+[.,][0-9]{2})"],
    "iva":      [r"(?:IVA|I\.V\.A\.|impuesto)[:\s]*(?:\d+\s*%\s*)?[:\s$]*([0-9]+[.,][0-9]{2})"],
    "total":    [r"(?:total\s*a\s*pagar|total\s*factura|importe\s*total|TOTAL)[:\s$]*([0-9]+[.,][0-9]{2})"],
    "moneda":   [r"\b(USD|ARS|EUR|MXN|CLP|COP|PEN|UYU)\b"],
    "condicion_pago": [
        r"(?:condici[oó]n|forma)\s+(?:de\s+)?pago[:\s]*([^\n\r]{3,40})",
        r"\b(contado|cr[eé]dito|30 d[ií]as|60 d[ií]as)\b",
    ],
}

COLUMNAS = {
    "archivo":        "Archivo PDF",
    "numero_factura": "N° Factura",
    "fecha_emision":  "Fecha Emisión",
    "emisor":         "Emisor / Proveedor",
    "cuit_emisor":    "CUIT / RUT",
    "cliente":        "Cliente",
    "subtotal":       "Subtotal",
    "iva":            "IVA",
    "total":          "Total",
    "moneda":         "Moneda",
    "condicion_pago": "Cond. de Pago",
    "paginas":        "Páginas",
    "error":          "Observaciones",
}


def limpiar_monto(t):
    t = t.strip()
    if re.match(r'^\d{1,3}(\.\d{3})*(,\d{2})$', t):
        return t.replace('.', '').replace(',', '.')
    if re.match(r'^\d{1,3}(,\d{3})*(\.\d{2})$', t):
        return t.replace(',', '')
    return t.replace(',', '.')


def extraer_campo(texto, campo):
    for patron in PATRONES.get(campo, []):
        m = re.search(patron, texto, re.IGNORECASE | re.MULTILINE)
        if m:
            return (m.group(1) if m.lastindex else m.group(0)).strip()
    return ""


def extraer_datos(pdf_bytes, nombre_archivo):
    datos = {k: "" for k in COLUMNAS}
    datos["archivo"] = nombre_archivo
    datos["paginas"] = 0
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            datos["paginas"] = len(pdf.pages)
            texto = "\n".join(p.extract_text() or "" for p in pdf.pages)
        if not texto.strip():
            datos["error"] = "PDF escaneado — sin texto extraíble"
            return datos
        for campo in PATRONES:
            datos[campo] = extraer_campo(texto, campo)
        for m in ["subtotal", "iva", "total"]:
            if datos[m]:
                datos[m] = limpiar_monto(datos[m])
    except Exception as e:
        datos["error"] = str(e)
    return datos


def generar_excel_bytes(registros):
    filas = [{COLUMNAS[k]: r.get(k, "") for k in COLUMNAS} for r in registros]
    df = pd.DataFrame(filas)

    buf = io.BytesIO()
    df.to_excel(buf, index=False, sheet_name="Facturas")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active

    hf  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    nf  = Font(name="Calibri", size=10)
    brd = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font  = hf
        cell.fill  = PatternFill("solid", fgColor="1F497D")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd

    col_obs = list(COLUMNAS.keys()).index("error") + 1
    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        tiene_error = bool(ws.cell(ri, col_obs).value)
        bg = "FFDAD6" if tiene_error else ("F2F2F2" if ri % 2 == 0 else "FFFFFF")
        for cell in row:
            cell.font   = nf
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center")
            cell.border = brd

    anchos = {
        "Archivo PDF": 28, "N° Factura": 18, "Fecha Emisión": 14,
        "Emisor / Proveedor": 32, "CUIT / RUT": 18, "Cliente": 28,
        "Subtotal": 14, "IVA": 12, "Total": 14, "Moneda": 10,
        "Cond. de Pago": 18, "Páginas": 10, "Observaciones": 32,
    }
    for i, name in enumerate(COLUMNAS.values(), 1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(name, 16)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ══════════════════════════════════════════════════════════════════════
#  INTERFAZ STREAMLIT
# ══════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Extractor de Facturas",
    page_icon="📄",
    layout="centered",
)

# ── CSS personalizado ──────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;600;700&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }

.header-band {
    background: linear-gradient(135deg, #1F497D 0%, #2E6DB4 100%);
    border-radius: 12px;
    padding: 28px 32px 22px;
    margin-bottom: 28px;
    color: white;
}
.header-band h1 { margin: 0; font-size: 1.8rem; font-weight: 700; }
.header-band p  { margin: 6px 0 0; opacity: .75; font-size: .95rem; }

.stat-card {
    background: white;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    padding: 16px 20px;
    text-align: center;
    box-shadow: 0 1px 4px rgba(0,0,0,.06);
}
.stat-num  { font-size: 2rem; font-weight: 700; color: #1F497D; }
.stat-label{ font-size: .8rem; color: #888; margin-top: 2px; }

.step-box {
    background: #1E3A5F;
    border-left: 4px solid #4A9EDB;
    border-radius: 0 8px 8px 0;
    padding: 14px 18px;
    margin-bottom: 10px;
    font-size: .93rem;
    color: #FFFFFF !important;
}
.step-num { font-weight: 700; color: #4A9EDB; margin-right: 8px; }
.step-box p, .step-box span, .step-box div { color: #FFFFFF !important; }

.ok-row   { color: #217346; }
.err-row  { color: #C0392B; }

div[data-testid="stFileUploader"] > label { display: none; }
</style>
""", unsafe_allow_html=True)

# ── Header ─────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-band">
  <h1>📄 Extractor de Facturas</h1>
  <p>Subí tus PDFs y descargá el Excel con todos los datos organizados — sin escribir nada a mano.</p>
</div>
""", unsafe_allow_html=True)

# ── Pasos ──────────────────────────────────────────────────────────────
with st.expander("ℹ️ ¿Cómo funciona? (3 pasos)", expanded=False):
    st.markdown("**1.** 📎 Adjunte su factura en formato PDF usando el botón de carga.")
    st.markdown("**2.** ⏳ Espere a que termine el proceso — la barra de progreso indica el avance.")
    st.markdown("**3.** ⬇️ Descargue su Excel con los datos de la factura usando el botón verde.")

st.divider()

# ── Upload ─────────────────────────────────────────────────────────────
st.markdown("#### 📂 Seleccioná tus facturas en PDF")
archivos = st.file_uploader(
    "Subir PDFs",
    type=["pdf"],
    accept_multiple_files=True,
    help="Podés seleccionar varios archivos a la vez.",
    label_visibility="collapsed",
)

if archivos:
    st.success(f"✅ {len(archivos)} archivo{'s' if len(archivos) > 1 else ''} cargado{'s' if len(archivos) > 1 else ''}.")

st.divider()

# ── Botón procesar ─────────────────────────────────────────────────────
col_btn, col_tip = st.columns([2, 3])
with col_btn:
    procesar = st.button("▶  Procesar facturas", type="primary",
                         disabled=not archivos, use_container_width=True)
with col_tip:
    if not archivos:
        st.caption("⬆️ Primero subí al menos un PDF.")

# ── Procesamiento ──────────────────────────────────────────────────────
if procesar and archivos:
    registros = []
    resultados_ui = []

    prog = st.progress(0, text="Iniciando…")
    total = len(archivos)

    for i, archivo in enumerate(archivos):
        prog.progress((i) / total, text=f"Procesando {archivo.name} ({i+1}/{total})…")
        datos = extraer_datos(archivo.read(), archivo.name)
        registros.append(datos)
        ok = not datos.get("error")
        resultados_ui.append((archivo.name, ok, datos.get("error", "")))

    prog.progress(1.0, text="¡Listo!")

    # ── Estadísticas ────────────────────────────────────────────────
    procesadas = sum(1 for _, ok, _ in resultados_ui if ok)
    con_error  = total - procesadas

    st.markdown("#### 📊 Resultado del proceso")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<div class="stat-card"><div class="stat-num">{total}</div>'
                    f'<div class="stat-label">Facturas procesadas</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-card"><div class="stat-num" style="color:#217346">{procesadas}</div>'
                    f'<div class="stat-label">Extraídas con éxito</div></div>', unsafe_allow_html=True)
    with c3:
        color_err = "#C0392B" if con_error else "#217346"
        st.markdown(f'<div class="stat-card"><div class="stat-num" style="color:{color_err}">{con_error}</div>'
                    f'<div class="stat-label">Con advertencias</div></div>', unsafe_allow_html=True)

    # ── Detalle por archivo ──────────────────────────────────────────
    with st.expander("Ver detalle por archivo"):
        for nombre, ok, err in resultados_ui:
            icono = "✅" if ok else "⚠️"
            nota  = "" if ok else f" — {err}"
            st.markdown(f"{icono} `{nombre}`{nota}")

    # ── Preview de datos ─────────────────────────────────────────────
    st.markdown("#### 👀 Vista previa del Excel")
    filas_preview = [{COLUMNAS[k]: r.get(k, "") for k in COLUMNAS} for r in registros]
    df_preview = pd.DataFrame(filas_preview)
    st.dataframe(df_preview, use_container_width=True, height=280)

    # ── Descarga ─────────────────────────────────────────────────────
    st.divider()
    excel_bytes = generar_excel_bytes(registros)
    nombre_excel = f"facturas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    st.download_button(
        label="⬇️  Descargar Excel",
        data=excel_bytes,
        file_name=nombre_excel,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
    st.caption(f"El archivo se llama `{nombre_excel}` y se guarda en tu carpeta de Descargas.")

# ── Footer ─────────────────────────────────────────────────────────────
st.markdown("---")
st.caption("💡 Funciona mejor con facturas electrónicas (AFIP, SAT, etc.). "
           "Los PDFs escaneados pueden no extraer datos correctamente.")
