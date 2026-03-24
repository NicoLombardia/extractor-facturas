"""
Extractor de Documentación Contable — La Nación
App web Streamlit — diseño minimalista corporativo.
"""

import io
import re
import base64
from pathlib import Path
from datetime import datetime

import pdfplumber
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
#  LOGO EMBEBIDO
# ══════════════════════════════════════════════════════════════════════

def get_logo_b64():
    for ruta in ["La_Nacion_Logo.png", "/mnt/user-data/uploads/La_Nacion_Logo.png"]:
        try:
            with open(ruta, "rb") as f:
                return base64.b64encode(f.read()).decode()
        except Exception:
            pass
    return ""


# ══════════════════════════════════════════════════════════════════════
#  EXTRACCIÓN OPTIMIZADA PARA FACTURAS DE LA NACIÓN
# ══════════════════════════════════════════════════════════════════════

def extraer_emisor(texto):
    lineas = [l.strip() for l in texto.split('\n') if l.strip()]

    # Razón social explícita (HandyWay style)
    m = re.search(r'[Rr]az[oó]n\s+social[:\s]+([^\n\r]{3,60})', texto)
    if m:
        val = m.group(1).strip()
        if val and 'LA NACION' not in val.upper() and 'NACION' not in val.upper():
            return val

    # Empresa en primeras líneas con S.A., CARGO, TRANSPORTES, etc.
    keywords = ['S.A.', 'SA ', ' SA\n', 'S.R.L.', 'SRL', 'CARGO', 'TRANSPORTES', 'SERVICIOS']
    for linea in lineas[:20]:
        tiene_kw = any(kw in linea.upper() for kw in keywords)
        if tiene_kw and 'LA NACION' not in linea.upper() and len(linea) > 5:
            return linea.strip()

    return ""


def extraer_fecha(texto):
    # "Fecha: DD/MM/YYYY"
    m = re.search(r'[Ff]echa[:\s]+(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})', texto)
    if m:
        return m.group(1).strip()

    # Aerolíneas: "06 03 2026" (tres bloques separados en el encabezado)
    m = re.search(r'\b(\d{2})\s+(\d{2})\s+(20\d{2})\b', texto)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"

    # Genérico DD/MM/YYYY
    m = re.search(r'\b(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](20\d{2})\b', texto)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"

    return ""


def parsear_monto(t):
    try:
        t = t.strip().replace(' ', '')
        if re.match(r'^\d{1,3}(\.\d{3})+(,\d{1,2})$', t):
            return float(t.replace('.', '').replace(',', '.'))
        if re.match(r'^\d{1,3}(,\d{3})+(\.\d{1,2})$', t):
            return float(t.replace(',', ''))
        if re.match(r'^\d+(,\d{1,2})$', t):
            return float(t.replace(',', '.'))
        return float(t.replace(',', ''))
    except Exception:
        return None


def formatear_monto(valor):
    try:
        partes = f"{valor:,.2f}".split('.')
        entero = partes[0].replace(',', '.')
        return f"$ {entero},{partes[1]}"
    except Exception:
        return str(valor)


def extraer_importe(texto):
    patrones = [
        r'[Ii]mporte\s+[Tt]otal\s+\$?\s*([\d\.,]+)',
        r'TOTAL\s+EN\s+PESOS\s+([\d\.,]+)',
        r'\bTOTAL\s+\$?\s*([\d\.,]+)',
        r'[Tt]otal\s*\$\s*([\d\.,]+)',
    ]
    candidatos = []
    for pat in patrones:
        for m in re.finditer(pat, texto):
            val = parsear_monto(m.group(1))
            if val and val > 100:
                candidatos.append(val)

    if not candidatos:
        return ""
    return formatear_monto(max(candidatos))


def extraer_numero_factura(texto):
    patrones = [
        r'[Nn]ro\.?:?\s*(\d{4,5}-\d{5,10})',
        r'[Cc]omprob\.?\s*[Nn]º?:?\s*(\d{4}-\d{5,10})',
        r'FACTURA\s*[:\s]*(\d{4}-\d{5,10})',
        r'(\d{4}-\d{6,10})',
    ]
    for pat in patrones:
        m = re.search(pat, texto)
        if m:
            return m.group(1).strip()
    return ""


def extraer_cuit_emisor(texto):
    for m in re.finditer(r'CUIT[:\s#°Nº]*(\d{2}[-\s]?\d{8}[-\s]?\d)', texto):
        cuit = m.group(1).strip()
        if '50008962' not in cuit:
            return cuit
    return ""


def extraer_datos(pdf_bytes, nombre_archivo):
    resultado = {
        "archivo": nombre_archivo,
        "emisor": "",
        "fecha_emision": "",
        "importe_total": "",
        "numero_factura": "",
        "cuit_emisor": "",
        "error": "",
    }
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            texto = "\n".join(p.extract_text() or "" for p in pdf.pages[:2])
        if not texto.strip():
            resultado["error"] = "PDF escaneado — sin texto extraíble"
            return resultado
        resultado["emisor"]         = extraer_emisor(texto)
        resultado["fecha_emision"]  = extraer_fecha(texto)
        resultado["importe_total"]  = extraer_importe(texto)
        resultado["numero_factura"] = extraer_numero_factura(texto)
        resultado["cuit_emisor"]    = extraer_cuit_emisor(texto)
    except Exception as e:
        resultado["error"] = str(e)
    return resultado


# ══════════════════════════════════════════════════════════════════════
#  EXCEL
# ══════════════════════════════════════════════════════════════════════

COLUMNAS = {
    "archivo":        "Archivo",
    "emisor":         "Empresa / Emisor",
    "fecha_emision":  "Fecha de Emisión",
    "importe_total":  "Importe Total",
    "numero_factura": "N° Comprobante",
    "cuit_emisor":    "CUIT Emisor",
    "error":          "Observaciones",
}


def generar_excel_bytes(registros):
    filas = [{COLUMNAS[k]: r.get(k, "") for k in COLUMNAS} for r in registros]
    buf = io.BytesIO()
    pd.DataFrame(filas).to_excel(buf, index=False, sheet_name="Facturas")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active

    hf  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    nf  = Font(name="Calibri", size=10, color="1A1A2E")
    bf  = Font(name="Calibri", size=10, bold=True, color="1B3A6B")
    brd = Border(
        left=Side(style="thin", color="D0D7E2"),
        right=Side(style="thin", color="D0D7E2"),
        top=Side(style="thin", color="D0D7E2"),
        bottom=Side(style="thin", color="D0D7E2"),
    )

    for cell in ws[1]:
        cell.font      = hf
        cell.fill      = PatternFill("solid", fgColor="1B3A6B")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = brd

    col_obs = list(COLUMNAS.keys()).index("error") + 1
    col_imp = list(COLUMNAS.keys()).index("importe_total") + 1

    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        tiene_error = bool(ws.cell(ri, col_obs).value)
        bg = "FFDAD6" if tiene_error else ("F0F4FA" if ri % 2 == 0 else "FFFFFF")
        for cell in row:
            cell.font      = nf
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center")
            cell.border    = brd
        ws.cell(ri, col_imp).font = bf

    anchos = {
        "Archivo": 30, "Empresa / Emisor": 35, "Fecha de Emisión": 16,
        "Importe Total": 20, "N° Comprobante": 22, "CUIT Emisor": 18,
        "Observaciones": 30,
    }
    for i, name in enumerate(COLUMNAS.values(), 1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(name, 18)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ══════════════════════════════════════════════════════════════════════
#  UI — DISEÑO MINIMALISTA LA NACIÓN
# ══════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Extractor Contable · La Nación",
    page_icon="📋",
    layout="centered",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Merriweather:wght@400;700&family=Inter:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

.ln-wrap {
    max-width: 680px;
    margin: 0 auto;
    padding: 36px 0 60px;
    text-align: center;
}

.ln-title {
    font-family: 'Merriweather', Georgia, serif;
    font-size: 1.25rem;
    font-weight: 700;
    color: #1B3A6B;
    line-height: 1.5;
    margin: 0 0 4px;
}

.ln-subtitle {
    font-size: 0.82rem;
    color: #8A97B0;
    font-weight: 300;
    margin: 0 0 20px;
    letter-spacing: 0.03em;
}

.ln-logo-wrap {
    display: flex;
    justify-content: center;
    align-items: center;
    padding: 12px 0 24px;
}
.ln-logo-wrap img {
    max-height: 48px;
    max-width: 220px;
    object-fit: contain;
}

.ln-rule {
    border: none;
    border-top: 1px solid #E2E8F0;
    margin: 0 0 28px;
}

.ln-upload-lbl {
    font-size: 0.78rem;
    font-weight: 600;
    color: #4A5568;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    text-align: left;
    margin-bottom: 6px;
}

.ln-stat {
    background: white;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    padding: 14px 8px;
    text-align: center;
}
.ln-stat-n { font-size: 1.9rem; font-weight: 700; color: #1B3A6B; line-height: 1; }
.ln-stat-l { font-size: 0.68rem; color: #A0AABF; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.06em; }

div[data-testid="stFileUploader"] > label { display: none; }

div[data-testid="stButton"] > button {
    background-color: #1B3A6B !important;
    color: white !important;
    border: none !important;
    border-radius: 6px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.9rem !important;
    padding: 10px 24px !important;
    letter-spacing: 0.04em !important;
    transition: background 0.2s;
}
div[data-testid="stButton"] > button:hover {
    background-color: #254D8F !important;
}
div[data-testid="stDownloadButton"] > button {
    background-color: #1B3A6B !important;
    color: white !important;
    border: none !important;
    border-radius: 6px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
}

.ln-footer {
    text-align: center;
    font-size: 0.72rem;
    color: #B0BAD0;
    margin-top: 40px;
    letter-spacing: 0.03em;
}
</style>
""", unsafe_allow_html=True)

# ── Título ───────────────────────────────────────────────────────────
st.markdown("""
<div style="text-align:center; padding: 32px 0 0;">
  <p class="ln-title">Extractor de datos de documentación<br>contable de La Nación</p>
  <p class="ln-subtitle">Procesamiento automático de facturas y comprobantes</p>
</div>
""", unsafe_allow_html=True)

# ── Logo ─────────────────────────────────────────────────────────────
logo_b64 = get_logo_b64()
if logo_b64:
    st.markdown(f"""
    <div class="ln-logo-wrap">
      <img src="data:image/png;base64,{logo_b64}" alt="La Nación"/>
    </div>
    """, unsafe_allow_html=True)
else:
    st.markdown('<div style="height:24px"></div>', unsafe_allow_html=True)

st.markdown('<hr class="ln-rule">', unsafe_allow_html=True)

# ── Upload ───────────────────────────────────────────────────────────
st.markdown('<p class="ln-upload-lbl">📎 &nbsp; Adjunte su documento</p>', unsafe_allow_html=True)

archivos = st.file_uploader(
    "PDFs",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    help="Puede seleccionar múltiples archivos PDF a la vez.",
)

if archivos:
    n = len(archivos)
    st.caption(f"✔  {n} archivo{'s' if n > 1 else ''} seleccionado{'s' if n > 1 else ''}.")
else:
    st.caption("Formatos admitidos: PDF · Factura electrónica AFIP")

st.markdown("<br>", unsafe_allow_html=True)

# ── Botón ────────────────────────────────────────────────────────────
procesar = st.button(
    "Procesar documentos",
    disabled=not archivos,
    use_container_width=True,
)

# ── Procesamiento ────────────────────────────────────────────────────
if procesar and archivos:
    registros, resultados_ui = [], []
    total = len(archivos)
    prog  = st.progress(0, text="Iniciando...")

    for i, archivo in enumerate(archivos):
        prog.progress(i / total, text=f"Procesando {archivo.name}…")
        datos = extraer_datos(archivo.read(), archivo.name)
        registros.append(datos)
        ok = not datos.get("error")
        resultados_ui.append((archivo.name, ok, datos.get("error", ""), datos))

    prog.progress(1.0, text="Completado.")
    st.markdown("<br>", unsafe_allow_html=True)

    # Estadísticas
    procesadas = sum(1 for _, ok, _, _ in resultados_ui if ok)
    con_error  = total - procesadas

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<div class="ln-stat"><div class="ln-stat-n">{total}</div><div class="ln-stat-l">Documentos</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="ln-stat"><div class="ln-stat-n" style="color:#2D6A4F">{procesadas}</div><div class="ln-stat-l">Procesados</div></div>', unsafe_allow_html=True)
    with c3:
        col = "#C0392B" if con_error else "#2D6A4F"
        st.markdown(f'<div class="ln-stat"><div class="ln-stat-n" style="color:{col}">{con_error}</div><div class="ln-stat-l">Con advertencias</div></div>', unsafe_allow_html=True)

    # Tabla
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("**Datos extraídos**")

    filas = []
    for _, ok, err, d in resultados_ui:
        filas.append({
            "Archivo":          d["archivo"],
            "Empresa / Emisor": d["emisor"] or "—",
            "Fecha de Emisión": d["fecha_emision"] or "—",
            "Importe Total":    d["importe_total"] or "—",
            "N° Comprobante":   d["numero_factura"] or "—",
            "Observaciones":    err or "OK",
        })

    st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

    # Descarga
    st.markdown("<br>", unsafe_allow_html=True)
    excel_bytes  = generar_excel_bytes(registros)
    nombre_excel = f"LaNacion_facturas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    st.download_button(
        label="⬇  Descargar Excel",
        data=excel_bytes,
        file_name=nombre_excel,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ── Footer ───────────────────────────────────────────────────────────
st.markdown("""
<p class="ln-footer">La Nación &nbsp;·&nbsp; Documentación Contable &nbsp;·&nbsp; Uso interno</p>
""", unsafe_allow_html=True)
