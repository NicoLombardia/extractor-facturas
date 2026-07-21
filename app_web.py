import io, re, base64
from datetime import datetime

import pdfplumber
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
#  LOGO
# ══════════════════════════════════════════════════════════════════════

def get_logo_b64():
    for ruta in ["La_Nacion_Logo.png", "/mnt/user-data/uploads/La_Nacion_Logo.png"]:
        try:
            with open(ruta, "rb") as f:
                return base64.b64encode(f.read()).decode()
        except Exception:
            pass
    return ""


# ══════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ══════════════════════════════════════════════════════════════════════

def limpiar_monto(s):
    s = str(s).strip().replace('\xa0', '')
    # Formato argentino: 1.234.567,89 → quitar puntos, cambiar coma
    if re.search(r'\d\.\d{3}', s):
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.')
    s = re.sub(r'[^\d\.]', '', s)
    try:
        return float(s)
    except Exception:
        return 0.0


MESES_ABREV = {
    '01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr',
    '05': 'May', '06': 'Jun', '07': 'Jul', '08': 'Aug',
    '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dec',
}

def mes_de_fecha(fecha_str):
    """Extrae número de mes de una fecha dd/mm/aa o dd/mm/yyyy."""
    m = re.search(r'(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})', str(fecha_str))
    if m:
        return int(m.group(2))
    return ""

def año_de_fecha(fecha_str):
    m = re.search(r'(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})', str(fecha_str))
    if m:
        y = m.group(3)
        return int(y) if len(y) == 4 else 2000 + int(y)
    return datetime.now().year

def normalizar_fecha(fecha_str):
    """Devuelve dd/mm/yyyy."""
    m = re.search(r'(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})', str(fecha_str))
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = '20' + y
        return f"{d.zfill(2)}/{mo.zfill(2)}/{y}"
    return str(fecha_str)


# ══════════════════════════════════════════════════════════════════════
#  DETECTAR PROVEEDOR
# ══════════════════════════════════════════════════════════════════════

def detectar_proveedor(texto_completo):
    t = texto_completo.upper()
    if 'AEROLINEAS ARGENTINAS' in t or '0458-' in t or 'AEROL' in t:
        return 'AEROLINEAS'
    if 'HANDYWAY' in t or 'HANDY WAY' in t:
        return 'HANDYWAY'
    if 'LIQUIDACION' in t and 'KGS:' in t:
        return 'HANDYWAY'
    if 'CRUZ DEL SUR' in t or '0514-' in t:
        return 'CRUZ DEL SUR'
    if '00048-' in t:
        return 'HANDYWAY'
    return 'DESCONOCIDO'


# ══════════════════════════════════════════════════════════════════════
#  PARSER: AEROLÍNEAS ARGENTINAS
#  Detalle en páginas con columnas:
#  NRO.GUIA | FECHA | KILOS | AFORADOS | TIPO-CARGA | ORIGEN | DESTINO | TRAMO | ITEM | IMPORTE
# ══════════════════════════════════════════════════════════════════════

# Mapeo de códigos IATA → nombre completo
IATA = {
    'AEP': 'BUENOS AIRES', 'BRC': 'BARILOCHE', 'JUJ': 'JUJUY', 'MDZ': 'MENDOZA',
    'NQN': 'NEUQUEN', 'SLA': 'SALTA', 'TUC': 'TUCUMAN', 'COR': 'CORDOBA',
    'BHI': 'BAHIA BLANCA', 'CRD': 'COMODORO RIVADAVIA', 'RGA': 'RIO GRANDE',
    'USH': 'USHUAIA', 'ROS': 'ROSARIO', 'PSS': 'POSADAS', 'RRL': 'TRELEW',
    'IGR': 'IGUAZU', 'SDR': 'SANTIAGO DEL ESTERO', 'FMA': 'FORMOSA',
    'SFE': 'SANTA FE', 'PRA': 'PARANA', 'IRJ': 'LA RIOJA', 'CTC': 'CATAMARCA',
    'LUQ': 'SAN LUIS', 'UAQ': 'SAN JUAN', 'RCL': 'RIO GALLEGOS',
    'RES': 'RESISTENCIA', 'VDM': 'VIEDMA', 'AFA': 'SAN RAFAEL',
    'RSA': 'SANTA ROSA', 'EQS': 'ESQUEL', 'CPC': 'CHAPELCO', 'MDQ': 'MAR DEL PLATA',
}

def iata(code):
    return IATA.get(code.strip().upper(), code.strip().upper())

def parse_aerolineas(pdf_bytes, nombre_archivo):
    filas = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        # Datos de cabecera desde página 1
        texto_p1 = pdf.pages[0].extract_text() or ""
        
        # Número de factura
        m = re.search(r'FACTURA\s*[:\s]*(\d{4}-\d{8})', texto_p1)
        nro_factura = m.group(1) if m else ""
        
        # Fecha factura (periodo al dd/mm/yy)
        m = re.search(r'Periodo\s+al[:\s]+(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})', texto_p1, re.IGNORECASE)
        if not m:
            m = re.search(r'(\d{2})\s+(\d{2})\s+(\d{4})', texto_p1)
            fecha_fac = f"{m.group(1)}/{m.group(2)}/{m.group(3)}" if m else ""
        else:
            fecha_fac = m.group(1)
        fecha_fac = normalizar_fecha(fecha_fac)
        
        # Línea de detalle: 044-XXXXXXXX DD-MM-YYYY kilos aforados TIPO ORI DST TRAMO ITEM importe
        # Patrón flexible que captura los tokens posicionales
        patron_guia = re.compile(
            r'044-\d+\s+'                              # nro guia
            r'(\d{1,2}-\d{2}-\d{4})\s+'               # fecha emision
            r'([\d,\.]+)\s+'                           # kilos reales (puede tener coma)
            r'(\d+)\s+'                                # kilos aforados (entero)
            r'(\w+)\s+'                                # tipo carga
            r'([A-Z]{3})\s+'                           # origen IATA
            r'([A-Z]{3})\s+'                           # destino IATA
            r'([A-Z]{3}-[A-Z]{3})\s+'                 # tramo
            r'(\d+)\s+'                                # item
            r'([\d\.,]+)'                              # importe
        )
        
        for page in pdf.pages[2:]:  # el detalle empieza en pág 3+
            texto = page.extract_text() or ""
            for m in patron_guia.finditer(texto):
                fecha_guia  = normalizar_fecha(m.group(1).replace('-', '/'))
                kilos       = int(m.group(3))
                origen      = iata(m.group(5))
                destino     = iata(m.group(6))
                tramo       = f"{origen} {destino}"
                importe     = limpiar_monto(m.group(9))
                
                filas.append({
                    'fecha':           fecha_fac,
                    'origen':          origen,
                    'destino':         destino,
                    'kilos_aforados':  kilos,
                    'clase':           'Prime',  # Aerolíneas siempre es Prime
                    'fac_total':       importe,
                    'mes':             mes_de_fecha(fecha_fac),
                    'proveedor':       'AEROLINEAS',
                    'tramo':           tramo,
                    'tipo_despacho':   'ENVIO',
                    'año':             año_de_fecha(fecha_fac),
                    'numero_factura':  nro_factura,
                    'archivo':         nombre_archivo,
                    'error':           '',
                })
    return filas


# ══════════════════════════════════════════════════════════════════════
#  PARSER: HANDYWAY CARGO (liquidaciones)
#  Cada línea: Fecha Guia# Origen Destino ... Kgs:XX.00 × Tarifa → importe
# ══════════════════════════════════════════════════════════════════════

def parse_handyway(pdf_bytes, nombre_archivo):
    """Liquidaciones Handyway — procesamiento línea por línea."""
    filas = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        lineas = []
        for p in pdf.pages:
            txt = p.extract_text() or ""
            lineas.extend(txt.split('\n'))

    texto_completo = '\n'.join(lineas)

    m = re.search(r'LIQUIDACION[:\s#]+(\d+)', texto_completo, re.IGNORECASE)
    nro_liq = m.group(1) if m else ""

    m = re.search(r'Fecha.Hora.{0,3}(\d{1,2}[/.]\d{2}[/.]\d{2,4})', texto_completo, re.IGNORECASE)
    fecha_liq = normalizar_fecha(m.group(1)) if m else ""

    patron_guia  = re.compile(r'^(\d{1,2}[/.]\d{2}[/.]\d{2})\s+(\d+)\s+([A-Z]{3})\s+([A-Z]{3})\s+')
    patron_kgs   = re.compile(r'(?:Kgs|PVol):([\d.]+)')
    patron_total = re.compile(r'=\s*\$\s*([\d.,]+)')

    i = 0
    while i < len(lineas):
        linea = lineas[i]
        m = patron_guia.match(linea)
        if m:
            origen  = iata(m.group(3))
            destino = iata(m.group(4))
            kilos = 0
            for j in range(i, min(i+4, len(lineas))):
                mk = patron_kgs.search(lineas[j])
                if mk:
                    kilos = int(round(float(mk.group(1))))
                    break
            importe = 0.0
            for j in range(i+1, min(i+5, len(lineas))):
                mt = patron_total.search(lineas[j])
                if mt:
                    importe = limpiar_monto(mt.group(1))
                    break
            if kilos > 0 or importe > 0:
                # Devolución si el origen no es Buenos Aires (AEP)
                tipo_d = 'DEVOLUCION' if m.group(3) != 'AEP' else 'ENVIO'
                filas.append({
                    'fecha':          fecha_liq,
                    'origen':         origen,
                    'destino':        destino,
                    'kilos_aforados': kilos,
                    'clase':          '',
                    'fac_total':      importe,
                    'mes':            mes_de_fecha(fecha_liq),
                    'proveedor':      'HANDYWAY',
                    'tramo':          f"{origen} {destino}",
                    'tipo_despacho':  tipo_d,
                    'año':            año_de_fecha(fecha_liq),
                    'numero_factura': nro_liq,
                    'archivo':        nombre_archivo,
                    'error':          '',
                })
        i += 1

    return filas
def parse_handyway_factura(pdf_bytes, nombre_archivo):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        texto = "\n".join(p.extract_text() or "" for p in pdf.pages)
    
    m = re.search(r'Nro\.\s*:\s*([\d\-]+)', texto)
    nro = m.group(1) if m else ""
    
    m = re.search(r'Fecha[:\s]+(\d{1,2}[\/\-\.]\d{2}[\/\-\.]\d{2,4})', texto, re.IGNORECASE)
    fecha = normalizar_fecha(m.group(1)) if m else ""
    
    m = re.search(r'Importe\s+Total\s*\$?\s*([\d\.,]+)', texto, re.IGNORECASE)
    total = limpiar_monto(m.group(1)) if m else 0.0
    
    m = re.search(r'Importe\s+neto\s+gravado\s*\$?\s*([\d\.,]+)', texto, re.IGNORECASE)
    neto = limpiar_monto(m.group(1)) if m else 0.0
    
    # Extraer nro liquidación del producto/servicio
    m2 = re.search(r'Liquidaci[oó]n\s*#?(\d+)', texto, re.IGNORECASE)
    ref_liq = m2.group(1) if m2 else ""
    
    return [{
        'fecha':          fecha,
        'origen':         'BUENOS AIRES',
        'destino':        '',
        'kilos_aforados': 0,
        'clase':          '',
        'fac_total':      neto,   # sin IVA, como en el modelo
        'mes':            mes_de_fecha(fecha),
        'proveedor':      'HANDYWAY',
        'tramo':          '',
        'tipo_despacho':  '',
        'año':            año_de_fecha(fecha),
        'numero_factura': nro,
        'archivo':        nombre_archivo,
        'error':          f'Liquidación #{ref_liq} — ver detalle adjunto' if ref_liq else '',
        '_advertencia':   True,
    }]


# ══════════════════════════════════════════════════════════════════════
#  PARSER: CRUZ DEL SUR — prioriza XLSX de detalle si está disponible
#  Si no, usa el texto del PDF
# ══════════════════════════════════════════════════════════════════════

def parse_cruz_del_sur_xlsx(xlsx_bytes, nombre_archivo, nro_factura, fecha_fac):
    """XLSX de detalle Cruz del Sur — usa Neto (sin IVA) como Fac Total por guía."""
    df = pd.read_excel(io.BytesIO(xlsx_bytes), header=0)

    filas = []
    for _, row in df.iterrows():
        kilos = 0
        try:
            v = row.get('KilogramosFacturados')
            if v is not None and str(v).strip() not in ('', 'nan'): kilos = int(float(v))
        except:
            try:
                v = row.get('Kilogramos')
                if v is not None and str(v).strip() not in ('', 'nan'): kilos = int(float(v))
            except:
                pass

        # Usar Neto (sin IVA) que coincide con el subtotal de la factura
        importe = 0.0
        for campo in ['Neto', 'TotalComprobante']:
            try:
                v = row.get(campo)
                if v is not None and str(v).strip() not in ('', 'nan'):
                    importe = float(v)
                    break
            except:
                pass

        destino_raw = str(row.get('DestinoLocalidad', '') or '').strip().upper()
        remitente   = str(row.get('Remitente', '') or '').strip().upper()

        # Si el remitente NO es La Nación, es una devolución (el origen es la provincia)
        es_la_nacion = 'LA NACION' in remitente or not remitente
        if es_la_nacion:
            origen_raw     = 'BUENOS AIRES'
            tipo_despacho  = 'ENVIO'
        else:
            # Devolución: origen es la localidad remitente, destino es BUE
            origen_raw    = destino_raw  # viene de provincia hacia BUE
            destino_raw   = 'BUENOS AIRES'
            tipo_despacho = 'DEVOLUCION'

        if not destino_raw or kilos == 0:
            continue

        tramo = f"{origen_raw} {destino_raw}"

        filas.append({
            'fecha':          fecha_fac,
            'origen':         origen_raw,
            'destino':        destino_raw,
            'kilos_aforados': kilos,
            'clase':          '',
            'fac_total':      importe,
            'mes':            mes_de_fecha(fecha_fac),
            'proveedor':      'CRUZ DEL SUR',
            'tramo':          tramo,
            'tipo_despacho':  tipo_despacho,
            'año':            año_de_fecha(fecha_fac),
            'numero_factura': nro_factura,
            'archivo':        nombre_archivo,
            'error':          '',
        })

    return filas



def parse_cruz_del_sur_pdf(pdf_bytes, nombre_archivo):
    """Parsea el PDF de Cruz del Sur (anexo con tabla Fecha|NIC|REMITO|FLETE|VARIOS|IMPORTE)."""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        texto = "\n".join(p.extract_text() or "" for p in pdf.pages)
    
    m = re.search(r'Comprob\.\s*N[°º]\s*:\s*([\d\-]+)', texto)
    nro = m.group(1) if m else ""
    
    m = re.search(r'Fecha[:\s]+(\d{1,2}[\/\-\.]\d{2}[\/\-\.]\d{4})', texto)
    fecha = normalizar_fecha(m.group(1)) if m else ""
    
    m = re.search(r'TOTAL\s*\$\s*([\d\.,]+)', texto)
    total = limpiar_monto(m.group(1)) if m else 0.0
    
    # Filas del anexo: Fecha NIC CartaDePorte REMITO flete 0.00 0.00 0.00 varios importe
    patron = re.compile(
        r'(\d{2}\/\d{2}\/\d{4})\s+'       # fecha
        r'(\d+)\s+'                         # NIC
        r'R-\d+-\d+\s+'                     # Carta de porte
        r'\d+-\d+\s+'                       # REMITO
        r'([\d,\.]+)\s+'                    # FLETE
        r'[\d\.,]+\s+[\d\.,]+\s+[\d\.,]+\s+([\d\.,]+)\s+'  # ACR CC GGRUA VARIOS
        r'([\d,\.]+)'                       # IMPORTE
    )
    
    filas = []
    for m2 in patron.finditer(texto):
        importe = limpiar_monto(m2.group(5))
        filas.append({
            'fecha':          fecha,
            'origen':         'BUENOS AIRES',
            'destino':        'RIO GRANDE',   # Cruz del Sur opera BUE-RGA típicamente
            'kilos_aforados': 0,              # No figura en el PDF, está en el XLSX
            'clase':          '',
            'fac_total':      importe,
            'mes':            mes_de_fecha(fecha),
            'proveedor':      'CRUZ DEL SUR',
            'tramo':          'BUENOS AIRES RIO GRANDE',
            'tipo_despacho':  'ENVIO',
            'año':            año_de_fecha(fecha),
            'numero_factura': nro,
            'archivo':        nombre_archivo,
            'error':          '',
        })
    
    if not filas:
        # Fallback: 1 fila con total
        filas.append({
            'fecha': fecha, 'origen': 'BUENOS AIRES', 'destino': '',
            'kilos_aforados': 0, 'clase': '', 'fac_total': total,
            'mes': mes_de_fecha(fecha), 'proveedor': 'CRUZ DEL SUR',
            'tramo': '', 'tipo_despacho': '', 'año': año_de_fecha(fecha),
            'numero_factura': nro, 'archivo': nombre_archivo, 'error': '',
        })
    
    return filas


# ══════════════════════════════════════════════════════════════════════
#  DISPATCHER PRINCIPAL
# ══════════════════════════════════════════════════════════════════════

def procesar_archivo(archivo_bytes, nombre, archivos_extra=None):
    """
    archivos_extra: dict {nombre_archivo: bytes} para archivos relacionados (xlsx de detalle).
    """
    try:
        # Detectar tipo
        nombre_up = nombre.upper()
        
        # ¿Es XLSX? (detalle Cruz del Sur)
        if nombre_up.endswith('.XLSX'):
            # Solo se procesa si viene acompañado del PDF correspondiente
            return []
        
        # Leer texto del PDF para detectar proveedor
        with pdfplumber.open(io.BytesIO(archivo_bytes)) as pdf:
            texto = "\n".join(p.extract_text() or "" for p in pdf.pages)
        
        proveedor = detectar_proveedor(texto)
        
        if proveedor == 'AEROLINEAS':
            return parse_aerolineas(archivo_bytes, nombre)
        
        elif proveedor == 'HANDYWAY':
            # ¿Es una liquidación o una factura formal?
            if 'LIQUIDACION' in texto.upper() and 'KGS:' in texto.upper():
                return parse_handyway(archivo_bytes, nombre)
            else:
                return parse_handyway_factura(archivo_bytes, nombre)
        
        elif proveedor == 'CRUZ DEL SUR':
            # Buscar si hay un XLSX de detalle con nombre similar
            nro = ""
            m = re.search(r'Comprob\.\s*N[°º]\s*:\s*([\d\-]+)', texto)
            if m:
                nro = m.group(1)
            m2 = re.search(r'Fecha[:\s]+(\d{1,2}[\/\-\.]\d{2}[\/\-\.]\d{4})', texto)
            fecha = normalizar_fecha(m2.group(1)) if m2 else ""
            
            if archivos_extra:
                for xnombre, xbytes in archivos_extra.items():
                    if xnombre.upper().endswith('.XLSX') and nro and nro.replace('-', '') in xnombre.replace('-', '').replace('_', ''):
                        return parse_cruz_del_sur_xlsx(xbytes, nombre, nro, fecha)
            
            return parse_cruz_del_sur_pdf(archivo_bytes, nombre)
        
        else:
            return [{'archivo': nombre, 'error': f'Proveedor no reconocido', 'fecha': '',
                     'origen': '', 'destino': '', 'kilos_aforados': 0, 'clase': '',
                     'fac_total': 0, 'mes': '', 'proveedor': '', 'tramo': '',
                     'tipo_despacho': '', 'año': '', 'numero_factura': ''}]
    
    except Exception as e:
        return [{'archivo': nombre, 'error': str(e), 'fecha': '',
                 'origen': '', 'destino': '', 'kilos_aforados': 0, 'clase': '',
                 'fac_total': 0, 'mes': '', 'proveedor': '', 'tramo': '',
                 'tipo_despacho': '', 'año': '', 'numero_factura': ''}]


# ══════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════

COLS = [
    ("Fecha",              18),
    ("Origen",             22),
    ("Destino",            22),
    ("kilos aforados",     16),
    ("Clase",              10),
    ("Fac Total",          16),
    ("Mes",                8),
    ("PROVEEDOR",          20),
    ("TRAMO",              36),
    ("TIPO DE DESPACHO",   20),
    ("AÑO",                8),
    ("NUMERO DE FACTURA",  22),
    ("PRECIO POR KILO",    18),
]

H_FILL  = PatternFill("solid", fgColor="1F4E79")
H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
S_FILL  = PatternFill("solid", fgColor="D6E4F0")
N_FILL  = PatternFill("solid", fgColor="FFFFFF")
BORDER  = Border(*[Side(style="thin", color="BDC3C7")]*0,
                 left=Side(style="thin", color="BDC3C7"),
                 right=Side(style="thin", color="BDC3C7"),
                 top=Side(style="thin", color="BDC3C7"),
                 bottom=Side(style="thin", color="BDC3C7"))
DF      = Font(name="Arial", size=10)
AC      = Alignment(horizontal="center", vertical="center")
AL      = Alignment(horizontal="left",   vertical="center")
AR      = Alignment(horizontal="right",  vertical="center")


def generar_excel(filas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Kilos"

    for ci, (nombre, ancho) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=nombre)
        c.font, c.fill, c.alignment, c.border = H_FONT, H_FILL, H_ALIGN, BORDER
        ws.column_dimensions[get_column_letter(ci)].width = ancho
    ws.row_dimensions[1].height = 30

    for ri, d in enumerate(filas, 2):
        fill = S_FILL if ri % 2 == 0 else N_FILL
        cF = get_column_letter(6)   # Fac Total
        cK = get_column_letter(4)   # kilos aforados

        vals = [
            d.get('fecha', ''),
            d.get('origen', ''),
            d.get('destino', ''),
            d.get('kilos_aforados') or '',
            d.get('clase', ''),
            d.get('fac_total') or '',
            d.get('mes', ''),
            d.get('proveedor', ''),
            d.get('tramo', ''),
            d.get('tipo_despacho', ''),
            d.get('año', ''),
            d.get('numero_factura', ''),
            None,  # PRECIO POR KILO — fórmula
        ]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci)
            c.fill, c.border, c.font = fill, BORDER, DF
            if ci == 13:
                c.value = f"=IFERROR({cF}{ri}/{cK}{ri},\"\")"
                c.alignment, c.number_format = AR, '#,##0.00'
            elif ci == 6:
                c.value = v
                c.alignment, c.number_format = AR, '#,##0.00'
            elif ci in (4, 7, 11):
                c.value = v
                c.alignment = AC
            else:
                c.value = v
                c.alignment = AL

    # Fila de totales
    tr = len(filas) + 2
    TF = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TL = PatternFill("solid", fgColor="2E75B6")
    
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=tr, column=1).alignment = AL

    for ci, col_l in [(4, get_column_letter(4)), (6, get_column_letter(6))]:
        c = ws.cell(row=tr, column=ci, value=f"=SUM({col_l}2:{col_l}{tr-1})")
        c.font, c.fill, c.border = TF, TL, BORDER
        c.alignment = AR if ci == 6 else AC
        if ci == 6: c.number_format = '#,##0.00'

    c13 = ws.cell(row=tr, column=13,
                  value=f"=IFERROR({get_column_letter(6)}{tr}/{get_column_letter(4)}{tr},\"\")")
    c13.font, c13.fill, c13.border = TF, TL, BORDER
    c13.alignment, c13.number_format = AR, '#,##0.00'

    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ══════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Extractor Contable · La Nación", page_icon="📄", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif!important;background-color:#F4F7FB!important;}
.ln-title{font-size:1.55rem;font-weight:600;color:#1F4E79;line-height:1.35;}
.ln-subtitle{font-size:.9rem;color:#6B7A99;margin-top:4px;}
.ln-logo-wrap{display:flex;justify-content:center;margin:18px 0 10px;}
.ln-logo-wrap img{height:48px;object-fit:contain;}
hr.ln-rule{border:none;border-top:1.5px solid #D0DAF0;margin:18px 0 22px;}
.ln-upload-lbl{font-size:.88rem;font-weight:500;color:#1F4E79;margin-bottom:4px;}
div[data-testid="stButton"]>button{background-color:#1F4E79!important;color:white!important;
  border:none!important;border-radius:6px!important;font-weight:500!important;
  font-size:.9rem!important;padding:10px 24px!important;}
div[data-testid="stDownloadButton"]>button{background-color:#1B3A6B!important;color:white!important;
  border:none!important;border-radius:6px!important;font-weight:500!important;}
.ln-footer{text-align:center;font-size:.72rem;color:#B0BAD0;margin-top:40px;}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div style="text-align:center;padding:32px 0 0;">
  <p class="ln-title">Extractor de datos de documentación<br>contable de La Nación</p>
  <p class="ln-subtitle">Procesamiento automático de facturas y comprobantes</p>
</div>""", unsafe_allow_html=True)

logo = get_logo_b64()
if logo:
    st.markdown(f'<div class="ln-logo-wrap"><img src="data:image/png;base64,{logo}" alt="La Nación"/></div>', unsafe_allow_html=True)
st.markdown('<hr class="ln-rule">', unsafe_allow_html=True)

st.markdown('<p class="ln-upload-lbl">📎 &nbsp; Adjunte sus documentos (PDF + XLSX de detalle si corresponde)</p>', unsafe_allow_html=True)

archivos = st.file_uploader("Archivos", type=["pdf","xlsx"],
                             accept_multiple_files=True, label_visibility="collapsed",
                             help="Subí los PDF de facturas. Para Cruz del Sur, también el XLSX de detalle con el mismo nombre de factura.")

if archivos:
    n = len(archivos)
    st.caption(f"✔  {n} archivo{'s' if n>1 else ''} seleccionado{'s' if n>1 else ''}.")
else:
    st.caption("Formatos: PDF · XLSX de detalle (Cruz del Sur)")

st.markdown("<br>", unsafe_allow_html=True)
procesar = st.button("Procesar documentos", disabled=not archivos, use_container_width=True)

if procesar and archivos:
    # Separar PDFs y XLSXs
    pdfs  = {a.name: a.read() for a in archivos if not a.name.upper().endswith('.XLSX')}
    xlsxs = {a.name: a.read() for a in archivos if a.name.upper().endswith('.XLSX')}
    
    todas_filas = []
    resultados_ui = []
    total = len(pdfs)
    prog = st.progress(0, text="Iniciando...")
    
    for i, (nombre, bites) in enumerate(pdfs.items()):
        prog.progress(i / total, text=f"Procesando {nombre}…")
        filas = procesar_archivo(bites, nombre, archivos_extra=xlsxs)
        todas_filas.extend(filas)
        errores = [f for f in filas if f.get('error') and not f.get('_advertencia')]
        advertencias = [f for f in filas if f.get('_advertencia')]
        resultados_ui.append((nombre, filas, errores, advertencias))
    
    prog.progress(1.0, text="Completado.")
    st.markdown("---")
    st.subheader("Resultados")
    
    for nombre, filas, errores, advs in resultados_ui:
        filas_ok = [f for f in filas if not f.get('error') and not f.get('_advertencia')]
        proveedor = filas[0].get('proveedor','') if filas else ''
        nro = filas[0].get('numero_factura','') if filas else ''
        
        if errores:
            st.error(f"❌ **{nombre}** — {errores[0]['error']}")
        elif advs:
            st.warning(f"⚠️ **{nombre}** — {advs[0]['error']}")
        else:
            total_kilos = sum(f.get('kilos_aforados',0) or 0 for f in filas_ok)
            total_imp   = sum(f.get('fac_total',0) or 0 for f in filas_ok)
            st.success(f"✅ **{nombre}** — {proveedor} · {len(filas_ok)} líneas · {total_kilos} kg · $ {total_imp:,.2f}")
    
    filas_validas = [f for f in todas_filas if not f.get('error') and not f.get('_advertencia')]
    if filas_validas:
        st.markdown("---")
        excel_bytes = generar_excel(filas_validas)
        st.download_button(
            "⬇️  Descargar Excel",
            data=excel_bytes,
            file_name=f"reporte_kilos_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.markdown('<div class="ln-footer">La Nación · Documentación Contable · Uso interno</div>', unsafe_allow_html=True)
